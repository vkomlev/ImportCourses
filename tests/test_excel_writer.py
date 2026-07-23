"""Тесты для модуля записи в Excel"""

import os
import tempfile
from pathlib import Path
from src.excel_writer import ExcelWriter, write_courses_to_excel
from openpyxl import load_workbook


def _safe_unlink(path):
    """Удаление файла (на Windows файл может быть ещё занят)."""
    try:
        os.unlink(path)
    except OSError:
        pass


class TestExcelWriter:
    """Тесты для класса ExcelWriter"""

    def test_init(self):
        """Тест инициализации writer"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            writer = ExcelWriter(path)
            assert writer.output_path == path
            assert writer.workbook is not None
            assert writer.worksheet is not None
            writer.workbook.close()
        finally:
            _safe_unlink(path)

    def test_write_headers(self):
        """Тест записи заголовков"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            writer = ExcelWriter(path)
            writer._write_headers()
            assert writer.worksheet.cell(row=1, column=1).value == "course_uid"
            assert writer.worksheet.cell(row=1, column=2).value == "title"
            assert writer.worksheet.cell(row=1, column=1).font.bold is True
            writer.workbook.close()
        finally:
            _safe_unlink(path)

    def test_format_bool(self):
        """Тест форматирования булевых значений"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            writer = ExcelWriter(path)
            assert writer._format_bool(True) == "true"
            assert writer._format_bool(False) == "false"
            assert writer._format_bool("true") == "true"
            assert writer._format_bool("false") == "false"
            assert writer._format_bool("1") == "true"
            assert writer._format_bool("0") == "false"
            assert writer._format_bool("yes") == "true"
            assert writer._format_bool("no") == "false"
            assert writer._format_bool("да") == "true"
            assert writer._format_bool("нет") == "false"
            assert writer._format_bool(None) == "false"
            writer.workbook.close()
        finally:
            _safe_unlink(path)

    def test_write_course(self):
        """Тест записи курса"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            writer = ExcelWriter(path)
            writer._write_headers()
            course = {
                "course_uid": "TEST-001",
                "title": "Тестовый курс",
                "description": "Описание",
                "access_level": "manual_check",
                "parent_course_uid": "",
                "order_number": "",
                "required_courses_uid": "",
                "is_required": "false",
            }
            writer._write_course(course)
            assert writer.worksheet.cell(row=2, column=1).value == "TEST-001"
            assert writer.worksheet.cell(row=2, column=2).value == "Тестовый курс"
            assert writer.worksheet.cell(row=2, column=4).value == "manual_check"
            writer.workbook.close()
        finally:
            _safe_unlink(path)

    def test_write_courses(self):
        """Тест записи списка курсов"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            courses = [
                {
                    "course_uid": "TEST-001",
                    "title": "Курс 1",
                    "description": "",
                    "access_level": "manual_check",
                    "parent_course_uid": "",
                    "order_number": "",
                    "required_courses_uid": "",
                    "is_required": "false",
                },
                {
                    "course_uid": "TEST-002",
                    "title": "Курс 2",
                    "description": "",
                    "access_level": "manual_check",
                    "parent_course_uid": "TEST-001",
                    "order_number": 1,
                    "required_courses_uid": "",
                    "is_required": "false",
                },
            ]
            writer = ExcelWriter(path)
            writer.write_courses(courses)
            assert os.path.exists(path)
            wb = load_workbook(path)
            ws = wb.active
            assert ws.cell(row=1, column=1).value == "course_uid"
            assert ws.cell(row=2, column=1).value == "TEST-001"
            assert ws.cell(row=3, column=1).value == "TEST-002"
            assert ws.cell(row=3, column=5).value == "TEST-001"
            assert ws.cell(row=3, column=6).value == 1
            wb.close()
        finally:
            _safe_unlink(path)


class TestWriteCoursesToExcelFunction:
    """Тесты для функции write_courses_to_excel"""

    def test_write_courses_to_excel(self):
        """Тест функции write_courses_to_excel"""
        courses = [
            {
                "course_uid": "TEST-001",
                "title": "Курс 1",
                "description": "",
                "access_level": "manual_check",
                "parent_course_uid": "",
                "order_number": "",
                "required_courses_uid": "",
                "is_required": "false",
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            path = tmp.name
        try:
            output_path = write_courses_to_excel(courses, path)
            assert output_path == path
            assert os.path.exists(output_path)
        finally:
            _safe_unlink(path)
