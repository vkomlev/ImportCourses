"""Запись данных курсов в XLSX файл"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from .config import DEFAULT_ACCESS_LEVEL, EXCEL_OUTPUT_DIR

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Класс для записи курсов в XLSX файл"""

    # Порядок колонок согласно мануалу
    COLUMNS = [
        "course_uid",
        "title",
        "description",
        "access_level",
        "parent_course_uid",
        "order_number",
        "required_courses_uid",
        "is_required",
    ]

    def __init__(self, output_path: Optional[str] = None):
        """
        Инициализация writer

        Args:
            output_path: Путь к выходному файлу (если None, генерируется автоматически)
        """
        self.output_path = output_path or self._generate_filename()
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Courses"

    def _generate_filename(self) -> str:
        """
        Генерация имени файла с timestamp

        Returns:
            Путь к файлу
        """
        output_dir = Path(EXCEL_OUTPUT_DIR)
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"courses_{timestamp}.xlsx"
        return str(output_dir / filename)

    def write_courses(self, courses: List[Dict[str, Any]]):
        """
        Запись списка курсов в Excel

        Args:
            courses: Список словарей с данными курсов
        """
        # Записываем заголовки
        self._write_headers()

        # Записываем данные курсов
        for course in courses:
            self._write_course(course)

        # Сохраняем файл
        self.workbook.save(self.output_path)
        logger.info(f"Файл сохранен: {self.output_path}")

    def _write_headers(self):
        """Запись заголовков колонок"""
        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = Font(bold=True)

    def _write_course(self, course: Dict[str, Any]):
        """
        Запись одного курса в Excel

        Args:
            course: Словарь с данными курса
        """
        row = self.worksheet.max_row + 1

        for col_idx, column_name in enumerate(self.COLUMNS, start=1):
            value = course.get(column_name, "")

            # Обработка специальных значений
            if column_name == "is_required":
                value = self._format_bool(value)
            elif column_name == "description" and value is None:
                value = ""
            elif column_name == "parent_course_uid" and value is None:
                value = ""
            elif column_name == "order_number":
                # order_number должен быть числом или пустой строкой
                if value is None or value == "":
                    value = ""
                else:
                    # Преобразуем в число, если возможно
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        value = ""
            elif column_name == "required_courses_uid" and value is None:
                value = ""
            elif column_name == "access_level" and not value:
                value = DEFAULT_ACCESS_LEVEL

            cell = self.worksheet.cell(row=row, column=col_idx)
            cell.value = value

    def _format_bool(self, value: Any) -> str:
        """
        Форматирование булевого значения для Excel

        Args:
            value: Значение (bool, str, int, None)

        Returns:
            Строка "true" или "false"
        """
        if isinstance(value, bool):
            return "true" if value else "false"

        if isinstance(value, str):
            value_lower = value.lower().strip()
            if value_lower in ("true", "1", "yes", "да"):
                return "true"
            elif value_lower in ("false", "0", "no", "нет", ""):
                return "false"

        if isinstance(value, (int, float)):
            return "true" if value else "false"

        # По умолчанию false
        return "false"

    def get_output_path(self) -> str:
        """
        Получить путь к выходному файлу

        Returns:
            Путь к файлу
        """
        return self.output_path


def write_courses_to_excel(courses: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
    """
    Удобная функция для записи курсов в Excel

    Args:
        courses: Список словарей с данными курсов
        output_path: Путь к выходному файлу (опционально)

    Returns:
        Путь к созданному файлу
    """
    writer = ExcelWriter(output_path)
    writer.write_courses(courses)
    return writer.get_output_path()


# Колонки для импорта материалов (docs/materials-import-template.md, materials-api.md)
MATERIALS_COLUMNS = [
    "course_uid",
    "external_uid",
    "title",
    "type",
    "url",
    "description",
    "caption",
    "order_position",
    "is_active",
]


def write_materials_to_excel(
    materials: List[Dict[str, Any]], output_path: Optional[str] = None
) -> str:
    """
    Запись материалов в XLSX для импорта в LMS (лист Materials).

    Args:
        materials: Список словарей с полями course_uid, external_uid, title, type, url, ...
        output_path: Путь к файлу (если None — output/materials_<timestamp>.xlsx)

    Returns:
        Путь к созданному файлу
    """
    output_dir = Path(EXCEL_OUTPUT_DIR)
    output_dir.mkdir(exist_ok=True)
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(output_dir / f"materials_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    for col_idx, col_name in enumerate(MATERIALS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True)

    for row_idx, mat in enumerate(materials, start=2):
        for col_idx, col_name in enumerate(MATERIALS_COLUMNS, start=1):
            value = mat.get(col_name, "")
            if col_name == "order_position" and value != "":
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    value = ""
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(output_path)
    logger.info(f"Файл материалов сохранен: {output_path}")
    return output_path
